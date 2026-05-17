import streamlit as st
import pandas as pd
import csv
import os
import io
from datetime import datetime
import re
import difflib
from google import genai
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from PIL import Image
import json

# ==========================================
# 🎨 1. MASTER PAGE SETUP & PRO STYLING
# ==========================================
st.set_page_config(page_title="Tally Automation Suite Pro", page_icon="🏢", layout="wide")

st.markdown("""
<style>
    .main { background-color: #F8F9FA; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; }
    h1, h2, h3 { color: #1E3A8A; font-weight: 600; }
    .stButton>button {
        background: linear-gradient(135deg, #1E40AF 0%, #3B82F6 100%);
        color: white; border-radius: 8px; border: none;
        padding: 12px 24px; font-size: 16px; font-weight: bold;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        transition: all 0.3s ease;
    }
    .stButton>button:hover { transform: translateY(-2px); box-shadow: 0 6px 12px rgba(0, 0, 0, 0.15); }
    .credit-box {
        background-color: #EFF6FF; border-left: 5px solid #3B82F6;
        padding: 15px; border-radius: 8px; margin-top: 25px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
    .credit-box h4 { margin: 0; color: #1D4ED8; font-size: 16px; }
    .credit-box p { margin: 5px 0 0 0; color: #475569; font-size: 14px; }
    div[data-baseweb="notification"] { border-radius: 8px; }
</style>
""", unsafe_allow_html=True)

# Constants & Paths
SUSPENSE_LEDGER = "Suspense A/c"
LEDGER_CASH = "Cash"
LEDGER_CARD = "Card"       
LEDGER_NBH  = "NBH"         
LEDGER_ONLINE = "Online"
MY_BANK_LEDGER = "HDFC Bank a/c"

MASTER_COMMON_PATH = "master_common_saved.csv"
MASTER_ARCADE_PATH = "master_arcade_saved.csv"
MASTER_STAFF_PATH  = "master_staff_saved.csv"

# ==========================================
# 🛠️ HELPER FUNCTIONS
# ==========================================
def escape_xml(text): return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

def parse_amount(val):
    try:
        if val is None or val == '': return 0.0
        return float(str(val).replace(',', '').replace('"', '').strip() or 0.0)
    except ValueError: return 0.0

def get_ledger_xml(l_name, is_deb, amt):
    is_deem = "Yes" if is_deb else "No"
    t_amt = f"-{amt}" if is_deb else f"{amt}"
    return f"<ALLLEDGERENTRIES.LIST><LEDGERNAME>{l_name}</LEDGERNAME><ISDEEMEDPOSITIVE>{is_deem}</ISDEEMEDPOSITIVE><AMOUNT>{t_amt}</AMOUNT></ALLLEDGERENTRIES.LIST>"

def find_ledger_exact(flat_no, master_list):
    f_val = str(flat_no).strip()
    if not f_val or f_val.lower() == 'nan': return SUSPENSE_LEDGER
    try:
        num = int(f_val)
        pattern = r'\b0*' + str(num) + r'\b'
    except ValueError: pattern = r'\b' + re.escape(f_val) + r'\b'
    for m in master_list:
        if re.search(pattern, m.upper()): return m
    return SUSPENSE_LEDGER

def get_best_ledger_match(excel_name, tally_ledger_list):
    if not tally_ledger_list: return excel_name.upper()
    excel_name_upper = str(excel_name).strip().upper()
    tally_ledgers_upper = [str(l).strip().upper() for l in tally_ledger_list]
    matches = difflib.get_close_matches(excel_name_upper, tally_ledgers_upper, n=1, cutoff=0.6)
    if matches:
        for original_ledger in tally_ledger_list:
            if str(original_ledger).strip().upper() == matches[0]: return original_ledger
    return excel_name.upper()

def load_master(path):
    if os.path.exists(path):
        try:
            return pd.read_csv(path, encoding='latin1').iloc[:, 0].dropna().astype(str).str.strip().tolist()
        except: return []
    return []

# 🖼️ SMART IMAGE COMPRESSOR (To Save Tokens)
def compress_image_for_ai(img, max_size=1024):
    if img.mode != 'RGB': img = img.convert('RGB')
    img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
    return img

def call_ai_with_retry(image, prompt_text, api_keys):
    if not api_keys: return None, "❌ Settings mein kam se kam ek API Key paste karein."
    compressed_img = compress_image_for_ai(image)
    
    for i, k in enumerate(api_keys):
        try:
            client = genai.Client(api_key=k)
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=[compressed_img, prompt_text]
            )
            return response.text, None
        except Exception as e:
            if "429" in str(e) or "RESOURCE_EXHAUSTED" in str(e):
                st.toast(f"🔄 Key {i+1} limit exhausted, trying next...")
                continue
            return None, f"Key {i+1} Error: {str(e)}"
    return None, "❌ Saari keys ki limit khatam ho chuki hai!"

# ==========================================
# 🔒 2. PASSWORD LOCK SYSTEM
# ==========================================
def check_password():
    def password_entered():
        if st.session_state["password"] == "23051987":
            st.session_state["password_correct"] = True
            del st.session_state["password"] 
        else: st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state or not st.session_state["password_correct"]:
        st.markdown("<br><br><h2 style='text-align: center; color: #1E3A8A;'>🔒 Tally Automation Suite Locked</h2>", unsafe_allow_html=True)
        st.text_input("Enter Admin Password", type="password", on_change=password_entered, key="password")
        if "password_correct" in st.session_state and not st.session_state["password_correct"]: st.error("❌ Invalid Password!")
        return False
    return True

# ==========================================
# 🚀 3. MAIN SOFTWARE EXECUTION
# ==========================================
if check_password():

    st.sidebar.markdown("## ⚙️ Control Panel")
    app_mode = st.sidebar.radio(
        "Select Module:",
        [
            "📊 Daily Collection", 
            "🏦 HDFC Bank Statement", 
            "🏢 Flats Monthly Billing", 
            "🏪 Arcade Monthly Billing", 
            "👤 Staff Salary Engine",
            "📸 Smart Attendance Logger"
        ]
    )
    
    st.sidebar.markdown("---")
    
    with st.sidebar.expander("⚙️ Update Master Ledgers", expanded=False):
        st.info("💡 One-time upload. Saves securely in system.")
        if c_file := st.file_uploader("1. Common Master", type=['csv']):
            with open(MASTER_COMMON_PATH, "wb") as f: f.write(c_file.getbuffer())
            st.success("Common Master Updated!")
        if a_file := st.file_uploader("2. Arcade Master", type=['csv']):
            with open(MASTER_ARCADE_PATH, "wb") as f: f.write(a_file.getbuffer())
            st.success("Arcade Master Updated!")
        if s_file := st.file_uploader("3. Staff Master", type=['csv']):
            with open(MASTER_STAFF_PATH, "wb") as f: f.write(s_file.getbuffer())
            st.success("Staff Master Updated!")

    st.sidebar.markdown("""
        <div class="credit-box">
            <h4>👨‍💻 Developed & Managed By</h4>
            <p><b>Yogesh Sharma</b><br>
            <i>Automation & Tally ERP Expert</i><br>
            📞 +91 8882516300</p>
        </div>
    """, unsafe_allow_html=True)

    # =========================================================================
    # 📊 TOOL 1: DAILY COLLECTION
    # =========================================================================
    if app_mode == "📊 Daily Collection":
        st.markdown("<h1 style='text-align: center;'>📊 Daily Receipt Automation</h1>", unsafe_allow_html=True)
        st.markdown("---")
        
        # 🚀 JADOO: Dono lists ko ek sath load kar raha hai (Flats + Arcade)
        common_list = load_master(MASTER_COMMON_PATH)
        arcade_list = load_master(MASTER_ARCADE_PATH)
        master_list = common_list + arcade_list
        
        if not master_list: 
            st.warning("⚠️ Masters missing! Please update Common & Arcade Masters from sidebar.")
        else: 
            st.success(f"✔️ System Ready: {len(master_list)} Ledgers Loaded (Flats + Shops)")

        data_file = st.file_uploader("Upload Daily Collection (Excel/CSV)", type=['xlsx', 'xls', 'csv'])
        
        if data_file and master_list:
            if st.button("🚀 Process Data & Generate XML"):
                with st.spinner('Processing Tally Entries...'):
                    try:
                        raw_df = pd.read_csv(data_file, encoding='latin1', header=None) if data_file.name.lower().endswith('.csv') else pd.read_excel(data_file, header=None)
                        
                        # Added 'shop' in heading search
                        header_row_idx = next((idx for idx, row in raw_df.iterrows() if 'date' in " ".join(row.fillna('').astype(str).str.lower()) and ('flat' in " ".join(row.fillna('').astype(str).str.lower()) or 'shop' in " ".join(row.fillna('').astype(str).str.lower()) or 'name' in " ".join(row.fillna('').astype(str).str.lower()))), None)
                        
                        if header_row_idx is None: st.stop()
                            
                        df = raw_df.iloc[header_row_idx+1:].copy()
                        df.columns = raw_df.iloc[header_row_idx].fillna('').astype(str).str.strip().str.lower()
                        
                        date_col = next((c for c in df.columns if c == 'date' or ('date' in c and 'cheque' not in c)), None)
                        # Added check for 'shop' column
                        flat_col = next((c for c in df.columns if 'flat' in c or 'shop' in c), None)
                        remark_col = next((c for c in df.columns if 'remark' in c), None)
                        chq_no_col = next((c for c in df.columns if 'cheque' in c and 'no' in c), None)

                        col_cash = next((c for c in df.columns if c == 'cash'), None)
                        col_card = next((c for c in df.columns if c == 'card'), None)
                        col_nbh = next((c for c in df.columns if c == 'nbh'), None)
                        col_online = next((c for c in df.columns if 'cheque' in c and 'online' in c), None)

                        df = df.dropna(subset=[date_col])

                        xml_content = """<?xml version="1.0" encoding="utf-8"?>\n<ENVELOPE>\n<HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER>\n<BODY>\n<IMPORTDATA>\n<REQUESTDESC>\n<REPORTNAME>Vouchers</REPORTNAME>\n<STATICVARIABLES><SVCURRENTCOMPANY>##SVCURRENTCOMPANY</SVCURRENTCOMPANY></STATICVARIABLES>\n</REQUESTDESC>\n<REQUESTDATA>\n"""
                        success_count = 0
                        failed_names = []

                        for _, row in df.iterrows():
                            raw_date = str(row.get(date_col, '')).strip()
                            if not raw_date or raw_date.lower() == 'nan': continue
                            try: vch_date = pd.to_datetime(raw_date, dayfirst=True).strftime('%Y%m%d')
                            except: vch_date = raw_date.replace('-', '').replace('/', '')

                            flat_no = str(row.get(flat_col, '')).strip()
                            remarks = str(row.get(remark_col, '')).strip()
                            chq_no = str(row.get(chq_no_col, '')).strip()
                            if chq_no and chq_no.lower() != 'nan': remarks = f"{remarks} [Ref: {chq_no}]".strip()

                            detected_ledger = find_ledger_exact(flat_no, master_list)
                            if detected_ledger == SUSPENSE_LEDGER: failed_names.append(flat_no)

                            payment_map = [('CASH', col_cash, LEDGER_CASH), ('CARD', col_card, LEDGER_CARD), ('NBH', col_nbh, LEDGER_NBH), ('ONLINE', col_online, LEDGER_ONLINE)]

                            for mode_name, col_name, debit_ledger_name in payment_map:
                                if not col_name: continue
                                try: amount = abs(float(str(row.get(col_name, '0')).replace(',', '').strip() or 0))
                                except: amount = 0.0

                                if amount > 0:
                                    narration = escape_xml(f"Collection via {mode_name} {remarks}".strip())
                                    xml_content += f"<TALLYMESSAGE xmlns:UDF=\"TallyUDF\"><VOUCHER VCHTYPE=\"Receipt\" ACTION=\"Create\"><DATE>{vch_date}</DATE><VOUCHERTYPENAME>Receipt</VOUCHERTYPENAME><NARRATION>{narration}</NARRATION>{get_ledger_xml(escape_xml(detected_ledger), False, amount)}{get_ledger_xml(escape_xml(debit_ledger_name), True, amount)}</VOUCHER></TALLYMESSAGE>"
                                    success_count += 1

                        xml_content += """</REQUESTDATA>\n</IMPORTDATA>\n</BODY>\n</ENVELOPE>"""
                        st.success(f"✅ Kaam Pura Hua! Total {success_count} Receipts ban gayi hain.")
                        st.download_button("📥 Download Daily Collection XML", data=xml_content.encode('utf-8'), file_name="tally_daily_collections.xml", mime="application/xml")
                        
                        if failed_names:
                            txt_content = "⚠️ YE FLATS/SHOPS NAHI MILE (Suspense A/c):\n\n" + "\n".join([f"- {n}" for n in set(failed_names)])
                            st.download_button("⚠️ Download Suspense List", data=txt_content.encode('utf-8'), file_name="daily_suspense.txt", mime="text/plain")

                    except Exception as e: st.error(f"❌ Error: {e}")

    # =========================================================================
    # 🏦 TOOL 2: HDFC BANK STATEMENT
    # =========================================================================
    elif app_mode == "🏦 HDFC Bank Statement":
        st.markdown("<h1 style='text-align: center;'>🏦 HDFC Bank Reconciliation</h1>", unsafe_allow_html=True)
        st.markdown("---")
        
        master_list = load_master(MASTER_COMMON_PATH)
        if not master_list: st.warning("⚠️ Common Master missing! Please update from sidebar.")
        else: st.success(f"✔️ System Ready: {len(master_list)} Ledgers Loaded")

        data_file = st.file_uploader("Upload HDFC Statement", type=['xlsx', 'xls', 'csv'])
        
        if data_file and master_list:
            def find_best_ledger_hdfc(narration_text):
                text_upper = str(narration_text).upper().strip()
                for kw in ["CARD", "POS"]: 
                    if kw in text_upper: return LEDGER_CARD
                for kw in ["NOBROKER", "NBH", "SETTLEMENT"]: 
                    if kw in text_upper: return LEDGER_NBH
                for kw in ["VIVISH"]: 
                    if kw in text_upper: return "Other Site Settlement A/c"
                for fn in master_list:
                    if fn.upper() in text_upper: return fn 
                return SUSPENSE_LEDGER

            if st.button("🚀 Process Bank & Generate XML"):
                with st.spinner('Bank Data processing...'):
                    try:
                        raw_df = pd.read_csv(data_file, encoding='latin1', header=None) if data_file.name.lower().endswith('.csv') else pd.read_excel(data_file, header=None)
                        header_row_idx = next((idx for idx, row in raw_df.iterrows() if 'narration' in " ".join(row.fillna('').astype(str).str.lower()) and 'date' in " ".join(row.fillna('').astype(str).str.lower())), None)
                        
                        if header_row_idx is None: st.stop()
                            
                        df = raw_df.iloc[header_row_idx+1:].copy()
                        df.columns = raw_df.iloc[header_row_idx].fillna('').astype(str).str.strip().str.lower()
                        
                        date_col = next((c for c in df.columns if c == 'date'), None)
                        narration_col = next((c for c in df.columns if 'narration' in c), None)
                        chq_col = next((c for c in df.columns if 'chq' in c or 'ref' in c), None)
                        withdraw_col = next((c for c in df.columns if 'withdraw' in c), None)
                        deposit_col = next((c for c in df.columns if 'deposit' in c), None)

                        df = df.dropna(subset=[date_col])

                        xml_content = """<?xml version="1.0" encoding="utf-8"?>\n<ENVELOPE>\n<HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER>\n<BODY>\n<IMPORTDATA>\n<REQUESTDESC>\n<REPORTNAME>Vouchers</REPORTNAME>\n<STATICVARIABLES><SVCURRENTCOMPANY>##SVCURRENTCOMPANY</SVCURRENTCOMPANY></STATICVARIABLES>\n</REQUESTDESC>\n<REQUESTDATA>\n"""
                        success_count = 0
                        card_count = 0
                        failed_names = []

                        for _, row in df.iterrows():
                            raw_date = str(row.get(date_col, '')).strip()
                            if not raw_date or raw_date.lower() == 'nan': continue
                            try: vch_date = pd.to_datetime(raw_date, dayfirst=True).strftime('%Y%m%d')
                            except: vch_date = raw_date.replace('-', '').replace('/', '')

                            base_narration = str(row.get(narration_col, '')).strip()
                            inst_no = str(row.get(chq_col, '')).strip().replace('.0', '') if chq_col else ''
                            
                            try: withdraw_amt = float(str(row.get(withdraw_col, '0')).replace(',', '').strip() or 0)
                            except: withdraw_amt = 0.0
                            try: deposit_amt = float(str(row.get(deposit_col, '0')).replace(',', '').strip() or 0)
                            except: deposit_amt = 0.0

                            if deposit_amt == 0 and withdraw_amt == 0: continue

                            detected_ledger = find_best_ledger_hdfc(base_narration)
                            if detected_ledger == SUSPENSE_LEDGER: failed_names.append(base_narration)

                            if detected_ledger in ["Other Site Settlement A/c", LEDGER_NBH, LEDGER_CARD]:
                                vch_type = 'Contra'
                                if detected_ledger == LEDGER_CARD: card_count += 1
                            else:
                                vch_type = 'Receipt' if deposit_amt > 0 else 'Payment'

                            debit_ledger = escape_xml(MY_BANK_LEDGER if deposit_amt > 0 else detected_ledger)
                            credit_ledger = escape_xml(detected_ledger if deposit_amt > 0 else MY_BANK_LEDGER)
                            final_amount = deposit_amt if deposit_amt > 0 else withdraw_amt
                            narration_safe = escape_xml(f"[Ref: {inst_no}] {base_narration}" if inst_no and inst_no.lower() != 'nan' else base_narration)

                            xml_content += f"<TALLYMESSAGE xmlns:UDF=\"TallyUDF\"><VOUCHER VCHTYPE=\"{vch_type}\" ACTION=\"Create\"><DATE>{vch_date}</DATE><VOUCHERTYPENAME>{vch_type}</VOUCHERTYPENAME><NARRATION>{narration_safe}</NARRATION>{get_ledger_xml(debit_ledger, True, final_amount)}{get_ledger_xml(credit_ledger, False, final_amount)}</VOUCHER></TALLYMESSAGE>"
                            success_count += 1

                        xml_content += """</REQUESTDATA>\n</IMPORTDATA>\n</BODY>\n</ENVELOPE>"""
                        st.success(f"✅ Total {success_count} Vouchers ban gaye hain. (Card Entries: {card_count})")
                        st.download_button("📥 Download Bank XML", data=xml_content.encode('utf-8'), file_name="tally_bank.xml", mime="application/xml")
                        
                    except Exception as e: st.error(f"Error: {e}")

    # =========================================================================
    # 🏢 TOOL 3: FLATS MONTHLY BILLING
    # =========================================================================
    elif app_mode == "🏢 Flats Monthly Billing":
        st.markdown("<h1 style='text-align: center;'>🏢 Flats Billing Engine</h1>", unsafe_allow_html=True)
        st.markdown("---")
        
        master_list = load_master(MASTER_COMMON_PATH)
        if not master_list: st.warning("⚠️ Common Master missing! Sidebar se update karein.")
        else: st.success(f"✔️ System Ready: {len(master_list)} Ledgers Loaded")

        billing_file = st.file_uploader("Upload Monthly Readings (CSV)", type=['csv'])
        
        col1, col2 = st.columns(2)
        with col1:
            billing_month = st.text_input("Billing Month:", "May 2026")
            entry_date = st.date_input("Voucher Date")
            date_val = entry_date.strftime("%Y%m%d")
        with col2:
            days_in_month = st.number_input("Total Days?", min_value=1, max_value=31, value=31)
            days_old_rate = st.number_input("Purana Rate Din", value=0)
            days_new_rate = st.number_input("Naya Rate Din", value=31)

        if billing_file and master_list:
            if st.button("🚀 Generate Sales Bills XML"):
                with st.spinner('Processing Bills...'):
                    try:
                        decoded_file = billing_file.getvalue().decode('utf-8-sig').splitlines()
                        reader = csv.DictReader(decoded_file)
                        xml_content = '''<ENVELOPE><HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER><BODY><IMPORTDATA><REQUESTDESC><REPORTNAME>Vouchers</REPORTNAME></REQUESTDESC><REQUESTDATA>'''
                        entry_count = 0
                        failed_names = []
                        
                        for row in reader:
                            clean_row = {k.strip().lower(): str(v).strip() for k, v in row.items() if k is not None}
                            flat_no = clean_row.get('flat no.', clean_row.get('flat no', ''))
                            if not flat_no: continue
                                
                            party_ledger = find_ledger_exact(flat_no, master_list)
                            if party_ledger == SUSPENSE_LEDGER: failed_names.append(flat_no)
                            party_ledger = escape_xml(party_ledger)
                            
                            area = parse_amount(clean_row.get('area', 0))
                            if area <= 0: continue
                            
                            cam = round((area * 1.65 * 12 / 365) * days_in_month, 2)
                            sink = round((area * 0.05 * 12 / 365) * days_in_month, 2)
                            elec = round((area * 0.39 * 12 / 365) * days_in_month, 2)
                            gst = round((area * 0.26 * 12 / 365) * days_in_month, 2)
                            vend = round((30.00 * 12 / 365) * days_in_month, 2)
                            pb_kva = parse_amount(clean_row.get('power backup', 0))
                            ml_kva = parse_amount(clean_row.get('main load', 0))
                            
                            pb_fix = round((pb_kva * 100 * 12 / 365) * days_in_month, 2)
                            ml_fix = round(((ml_kva * 60 * 12 / 365) * days_old_rate) + ((ml_kva * 36.69 * 12 / 365) * days_new_rate), 2)
                            
                            m_units = max(0, parse_amount(clean_row.get('main curr', 0)) - parse_amount(clean_row.get('main prev', 0)))
                            b_units = max(0, parse_amount(clean_row.get('backup curr', 0)) - parse_amount(clean_row.get('backup prev', 0)))
                            
                            m_cons = round(m_units * 6.93, 2)
                            b_cons = max(0, round((b_units * 27.49) - 100, 2))
                            
                            total = round(cam + sink + elec + vend + gst + pb_fix + ml_fix + m_cons + b_cons, 2)
                            if total <= 0: continue

                            safe_narration = escape_xml(f"Bill for {billing_month} ({days_in_month} Days) | Area: {area} | Main: {int(m_units)}U | DG: {int(b_units)}U")

                            v_xml = f'''<TALLYMESSAGE xmlns:UDF="TallyUDF"><VOUCHER VCHTYPE="Sales" ACTION="Create"><DATE>{date_val}</DATE><VOUCHERTYPENAME>Sales</VOUCHERTYPENAME><NARRATION>{safe_narration}</NARRATION><ALLLEDGERENTRIES.LIST><LEDGERNAME>{party_ledger}</LEDGERNAME><ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE><AMOUNT>-{total}</AMOUNT></ALLLEDGERENTRIES.LIST>'''
                            
                            incomes = [("Common Area Maintenance", cam), ("Sinking Fund", sink), ("Common Area Electricity Charges", elec), ("Power Backup Charges", pb_fix), ("Main Load Charges", ml_fix), ("Main Load Consumption", m_cons), ("Power Backup Consumption", b_cons), ("Vending Charges", vend), ("GST Collection A/c", gst)]
                            
                            for l_name, amt in incomes:
                                if amt > 0: v_xml += f'''<ALLLEDGERENTRIES.LIST><LEDGERNAME>{escape_xml(l_name)}</LEDGERNAME><ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE><AMOUNT>{amt}</AMOUNT></ALLLEDGERENTRIES.LIST>'''
                            
                            v_xml += '''</VOUCHER></TALLYMESSAGE>\n'''
                            xml_content += v_xml
                            entry_count += 1
                            
                        xml_content += """</REQUESTDATA></IMPORTDATA></BODY></ENVELOPE>"""
                        st.success(f"✅ Total {entry_count} Flats generated!")
                        st.download_button("📥 Download Sales XML", data=xml_content.encode('utf-8'), file_name="tally_flats_sales.xml")
                        
                        if failed_names:
                            txt_content = "⚠️ YE FLATS NAHI MILE:\n" + "\n".join([f"- {n}" for n in set(failed_names)])
                            st.download_button("⚠️ Download Suspense List", data=txt_content.encode('utf-8'), file_name="flat_suspense.txt")
                    except Exception as e: st.error(f"Error: {e}")

    # =========================================================================
    # 🏪 TOOL 4: ARCADE MONTHLY BILLING
    # =========================================================================
    elif app_mode == "🏪 Arcade Monthly Billing":
        st.markdown("<h1 style='text-align: center;'>🏪 Arcade Billing Engine</h1>", unsafe_allow_html=True)
        st.markdown("---")
        
        master_list = load_master(MASTER_ARCADE_PATH)
        if not master_list: st.warning("⚠️ Arcade Master missing! Sidebar se update karein.")
        else: st.success(f"✔️ System Ready: {len(master_list)} Arcade Shops Loaded")

        billing_file = st.file_uploader("Upload Arcade Readings (CSV)", type=['csv'])
        
        col1, col2 = st.columns(2)
        with col1:
            billing_month = st.text_input("Month:", "May 2026")
            entry_date = st.date_input("Voucher Date")
            date_val = entry_date.strftime("%Y%m%d")
        with col2:
            days_in_month = st.number_input("Total Days:", value=31)

        if billing_file and master_list:
            if st.button("🚀 Generate ARCADE XML"):
                with st.spinner("Processing Arcade Bills..."):
                    try:
                        decoded_file = billing_file.getvalue().decode('utf-8-sig').splitlines()
                        reader = csv.DictReader(decoded_file)
                        xml_content = '''<ENVELOPE><HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER><BODY><IMPORTDATA><REQUESTDESC><REPORTNAME>Vouchers</REPORTNAME></REQUESTDESC><REQUESTDATA>'''
                        count = 0
                        failed_names = []

                        for row in reader:
                            c_row = {k.strip().lower(): str(v).strip() for k, v in row.items() if k is not None}
                            shop = c_row.get('shop no.', c_row.get('shop no', ''))
                            if not shop or shop == 'nan': continue
                            
                            party = find_ledger_exact(shop, master_list)
                            if party == SUSPENSE_LEDGER: failed_names.append(shop)
                            
                            m_prev = parse_amount(c_row.get('main prev', 0))
                            m_curr = parse_amount(c_row.get('main curr', 0))
                            b_prev = parse_amount(c_row.get('backup prev', 0))
                            b_curr = parse_amount(c_row.get('backup curr', 0))
                            
                            m_units = max(0, m_curr - m_prev)
                            b_units = max(0, b_curr - b_prev)

                            load = parse_amount(c_row.get('main load', 0))
                            cam = round(62.07 * days_in_month, 2)
                            m_fixed = round((load * 60 * 12 / 365) * days_in_month, 2)
                            pb_fixed = round((parse_amount(c_row.get('power backup', 0)) * 100 * 12 / 365) * days_in_month, 2)
                            vending = round((30.00 * 12 / 365) * days_in_month, 2)
                            
                            m_cons = round(m_units * 6.93, 2)
                            dg_cons = round(b_units * 27.49, 2) 
                            gst = parse_amount(c_row.get('gst', 0))
                            tass = parse_amount(c_row.get('tass', 0))
                            
                            total = round(cam + m_fixed + pb_fixed + vending + m_cons + dg_cons + gst + tass, 2) 
                            if total <= 0: continue
                            
                            final_narration = f"Arcade Bill {billing_month} | Shop: {shop} | Main: {int(m_units)}U | DG: {int(b_units)}U"
                            
                            xml_content += f'''<TALLYMESSAGE xmlns:UDF="TallyUDF"><VOUCHER VCHTYPE="Sales" ACTION="Create"><DATE>{date_val}</DATE><VOUCHERTYPENAME>Sales</VOUCHERTYPENAME><PARTYLEDGERNAME>{escape_xml(party)}</PARTYLEDGERNAME><NARRATION>{escape_xml(final_narration)}</NARRATION><ALLLEDGERENTRIES.LIST><LEDGERNAME>{escape_xml(party)}</LEDGERNAME><ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE><AMOUNT>-{total}</AMOUNT></ALLLEDGERENTRIES.LIST>'''
                            
                            incomes = [("Common Area Maintenance", cam), ("Main Load Charges", m_fixed), ("Power Backup Charges", pb_fixed), ("Vending Charges", vending), ("Main Load Consumption", m_cons), ("Power Backup Consumption", dg_cons), ("GST Collection A/c", gst)]
                            if tass > 0: incomes.append(("Common Area Electricity Charges", tass))

                            for l, a in incomes:
                                if a > 0: xml_content += f'''<ALLLEDGERENTRIES.LIST><LEDGERNAME>{escape_xml(l)}</LEDGERNAME><ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE><AMOUNT>{a}</AMOUNT></ALLLEDGERENTRIES.LIST>'''
                            
                            xml_content += '''</VOUCHER></TALLYMESSAGE>'''
                            count += 1

                        xml_content += """</REQUESTDATA></IMPORTDATA></BODY></ENVELOPE>"""
                        st.success(f"✅ {count} Arcade Bills ready!")
                        st.download_button("📥 Download Arcade XML", data=xml_content.encode('utf-8'), file_name="Arcade_Sales.xml")
                        
                        if failed_names:
                            txt_content = "⚠️ YE SHOPS NAHI MILI:\n" + "\n".join([f"- {n}" for n in set(failed_names)])
                            st.download_button("⚠️ Download Suspense List", data=txt_content.encode('utf-8'), file_name="arcade_suspense.txt")
                    except Exception as e: st.error(f"Error: {e}")

    # =========================================================================
    # 👤 TOOL 5: STAFF SALARY ENGINE
    # =========================================================================
    elif app_mode == "👤 Staff Salary Engine":
        st.markdown("<h1 style='text-align: center;'>👤 Smart Salary & Compliance</h1>", unsafe_allow_html=True)
        st.markdown("---")

        master_list = load_master(MASTER_STAFF_PATH)
        if not master_list: st.info("ℹ️ Staff Master not found. Auto-spell correction is OFF until updated from sidebar.")
        else: st.success(f"✔️ System Ready: {len(master_list)} Staff Ledgers Loaded (Auto Spell check ON)")

        salary_file = st.file_uploader("Upload Salary Sheet (Excel/CSV)", type=['xlsx', 'csv'])
        
        col1, col2 = st.columns(2)
        with col1:
            salary_month = st.text_input("Month Name:", "May 2026")
            days_in_month = st.number_input("Total Days in Month:", value=31)
            entry_date = st.date_input("Voucher Date")
            date_val = entry_date.strftime("%Y%m%d")
        with col2:
            exp_ledger = st.text_input("Salary Expense A/c (Dr):", "Site Staff Salary A/c")
            pf_ledger = st.text_input("PF Payable A/c (Cr):", "PF Payable A/c")
            esi_ledger = st.text_input("ESI Payable A/c (Cr):", "ESI Payable A/c")
            pen_ledger = st.text_input("Staff Penalty A/c (Cr):", "Staff Penalty/Fine A/c")

        if salary_file:
            if st.button("🚀 Generate Salary XML"):
                with st.spinner("Processing Entries..."):
                    try:
                        df = pd.read_excel(salary_file) if salary_file.name.endswith('.xlsx') else pd.read_csv(salary_file)
                        df.columns = df.columns.str.strip().str.lower()
                        
                        xml_content = '''<ENVELOPE><HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER><BODY><IMPORTDATA><REQUESTDESC><REPORTNAME>Vouchers</REPORTNAME></REQUESTDESC><REQUESTDATA>'''
                        count = 0
                        
                        for _, row in df.iterrows():
                            raw_emp_name = str(row.get('name', '')).strip()
                            actual_days = parse_amount(row.get('actual present days', 0))
                            if not raw_emp_name or actual_days <= 0: continue
                            
                            emp_ledger_name = get_best_ledger_match(raw_emp_name, master_list) if master_list else raw_emp_name.upper()
                            
                            fixed_basic = parse_amount(row.get('fixed basic', 0))
                            fixed_other = parse_amount(row.get('fixed other', 0))
                            att_perc = parse_amount(row.get('att % (multiplier)', 1.0)) or 1.0
                            
                            comp_days = round(actual_days / att_perc, 2)
                            earned_basic = round((fixed_basic / days_in_month) * comp_days, 0)
                            earned_other = round((fixed_other / days_in_month) * comp_days, 0)
                            gross = earned_basic + earned_other
                            
                            pf = round(earned_basic * 0.12, 0) if str(row.get('pf applicable?', '')).lower() == 'yes' else 0
                            esi = round(gross * 0.0075, 0) if str(row.get('esi applicable?', '')).lower() == 'yes' else 0
                            penalty = parse_amount(row.get('penalty/fine', 0))
                            advance = parse_amount(row.get('advance deduction', 0))
                            
                            net_to_ledger = gross - pf - esi - penalty - advance
                            if net_to_ledger <= 0: continue
                            
                            xml_content += f'''<TALLYMESSAGE xmlns:UDF="TallyUDF"><VOUCHER VCHTYPE="Journal" ACTION="Create"><DATE>{date_val}</DATE><VOUCHERTYPENAME>Journal</VOUCHERTYPENAME><NARRATION>Salary for {escape_xml(raw_emp_name)} | Days: {actual_days}</NARRATION>'''
                            xml_content += f'''<ALLLEDGERENTRIES.LIST><LEDGERNAME>{escape_xml(exp_ledger)}</LEDGERNAME><ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE><AMOUNT>-{gross}</AMOUNT></ALLLEDGERENTRIES.LIST>'''
                            if pf > 0: xml_content += f'''<ALLLEDGERENTRIES.LIST><LEDGERNAME>{escape_xml(pf_ledger)}</LEDGERNAME><ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE><AMOUNT>{pf}</AMOUNT></ALLLEDGERENTRIES.LIST>'''
                            if esi > 0: xml_content += f'''<ALLLEDGERENTRIES.LIST><LEDGERNAME>{escape_xml(esi_ledger)}</LEDGERNAME><ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE><AMOUNT>{esi}</AMOUNT></ALLLEDGERENTRIES.LIST>'''
                            if penalty > 0: xml_content += f'''<ALLLEDGERENTRIES.LIST><LEDGERNAME>{escape_xml(pen_ledger)}</LEDGERNAME><ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE><AMOUNT>{penalty}</AMOUNT></ALLLEDGERENTRIES.LIST>'''
                            xml_content += f'''<ALLLEDGERENTRIES.LIST><LEDGERNAME>{escape_xml(emp_ledger_name)}</LEDGERNAME><ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE><AMOUNT>{net_to_ledger}</AMOUNT></ALLLEDGERENTRIES.LIST></VOUCHER></TALLYMESSAGE>'''
                            count += 1

                        xml_content += '''</REQUESTDATA></IMPORTDATA></BODY></ENVELOPE>'''
                        st.success(f"✅ {count} Salary Vouchers Ready!")
                        st.download_button("📥 Download Salary XML", data=xml_content.encode('utf-8'), file_name="Salary_Vouchers.xml")
                    except Exception as e: st.error(f"Error: {e}")

    # =========================================================================
    # 📸 TOOL 6: SMART ATTENDANCE LOGGER
    # =========================================================================
    elif app_mode == "📸 Smart Attendance Logger":
        st.markdown("<h1 style='text-align: center;'>📝 Smart Attendance Logger</h1>", unsafe_allow_html=True)
        st.markdown("---")
        
        with st.expander("🔑 AI API Settings", expanded=True):
            col_k1, col_k2 = st.columns(2)
            with col_k1: key1 = st.text_input("Google Gemini API Key 1", type="password")
            with col_k2: key2 = st.text_input("API Key 2 (Optional Backup)", type="password")
            api_keys = [k.strip() for k in [key1, key2] if k.strip()]

        col1, col2 = st.columns([1, 2])
        with col1:
            attendance_excel = st.file_uploader("Upload Working Attendance Sheet (Excel)", type=["xlsx"])
            uploaded_image = st.file_uploader("Upload Attendance Register Scan", type=["jpg", "jpeg", "png"])
            
            st.markdown("### Settings")
            sheet_name = st.text_input("Excel Sheet Name", value="Apr.26")
            date_val = st.number_input("Aaj ki Date (1-31)", min_value=1, max_value=31, value=1)
            start_col_str = st.text_input("1 Tareekh ka 'IN' Column", value="F")

            try:
                start_idx = column_index_from_string(start_col_str.strip().upper())
                target_in_idx = start_idx + ((date_val - 1) * 2)
                target_out_idx = target_in_idx + 1
                col_in_let = get_column_letter(target_in_idx)
                col_out_let = get_column_letter(target_out_idx)
                st.success(f"🎯 Target: Columns {col_in_let} (IN) & {col_out_let} (OUT)")
            except:
                st.error("❌ Column letter sahi nahi hai.")
                st.stop()

        with col2:
            if uploaded_image and attendance_excel:
                img = Image.open(uploaded_image)
                st.image(img, use_container_width=True, caption="Uploaded Register (Will be compressed for AI)")

                if st.button("🚀 AI Scan & Update Excel", use_container_width=True):
                    with st.spinner("AI Handwriting aur AM/PM format check kar raha hai..."):
                        prompt = """
                        Scan this sheet. Extract Staff Name, IN, OUT.
                        STRICT RULES:
                        1. TIME FORMAT: Always include AM or PM. Do not give only numbers.
                        2. If handwritten 'N' or '(N)' is found, set "shift": "N".
                        3. If 'D+N' is found, set "shift": "DN".
                        4. Otherwise set "shift": "Day".
                        Return strictly JSON: {"NAME": {"in": "HH:MM AM/PM", "out": "HH:MM AM/PM", "shift": "Day/N/DN"}}
                        """
                        raw_output, err = call_ai_with_retry(img, prompt, api_keys)
                        
                        if err: st.error(err)
                        elif raw_output:
                            try:
                                clean_text = raw_output.strip()
                                if "```json" in clean_text: 
                                    clean_text = clean_text.split("```json")[1].split("```")[0].strip()
                                elif "```" in clean_text: 
                                    clean_text = clean_text.split("```")[1].split("```")[0].strip()
                                extracted_data = json.loads(clean_text)
                                st.success("✅ AI ne data padh liya!")
                                
                                wb = openpyxl.load_workbook(attendance_excel)
                                if sheet_name not in wb.sheetnames:
                                    st.error(f"Sheet '{sheet_name}' nahi mili!")
                                    st.stop()
                                    
                                ws = wb[sheet_name]
                                name_col, head_row = None, None
                                for r in range(1, 15):
                                    for c in range(1, 15):
                                        if str(ws.cell(row=r, column=c).value).strip().upper() == 'NAME':
                                            name_col, head_row = c, r
                                            break
                                    if name_col: break
                                    
                                if not name_col:
                                    st.error("Excel mein 'NAME' column nahi mila!")
                                    st.stop()

                                excel_names = {str(ws.cell(row=r, column=name_col).value).strip().lower(): r 
                                               for r in range(head_row + 1, ws.max_row + 1) if ws.cell(row=r, column=name_col).value}
                                
                                updated = 0
                                for ai_name, info in extracted_data.items():
                                    low_name = ai_name.strip().lower()
                                    row_idx = None
                                    
                                    if low_name in excel_names: row_idx = excel_names[low_name]
                                    else:
                                        matches = difflib.get_close_matches(low_name, list(excel_names.keys()), n=1, cutoff=0.7)
                                        if matches: row_idx = excel_names[matches[0]]
                                        else:
                                            first = low_name.split()[0] if low_name.split() else ""
                                            for ex_n, r_val in excel_names.items():
                                                if first and ex_n.startswith(first):
                                                    row_idx = r_val
                                                    break
                                                    
                                    if row_idx:
                                        shift = info.get('shift', 'Day')
                                        in_time = str(info.get('in', 'A')).upper()
                                        out_time = str(info.get('out', 'A')).upper()

                                        if ":" in in_time and "AM" not in in_time and "PM" not in in_time:
                                            in_time += " AM" if shift == "Day" else " PM"
                                        if ":" in out_time and "AM" not in out_time and "PM" not in out_time:
                                            out_time += " PM" if shift == "Day" else " AM"

                                        if shift == "DN":
                                            ws.cell(row=row_idx, column=target_in_idx).value = "D/N"
                                            ws.cell(row=row_idx, column=target_out_idx).value = "D/N"
                                        elif shift == "N":
                                            ws.cell(row=row_idx, column=target_in_idx).value = out_time
                                            ws.cell(row=row_idx, column=target_out_idx).value = in_time
                                        else:
                                            ws.cell(row=row_idx, column=target_in_idx).value = in_time
                                            ws.cell(row=row_idx, column=target_out_idx).value = out_time
                                        updated += 1
                                        
                                output = io.BytesIO()
                                wb.save(output)
                                output.seek(0)
                                
                                st.success(f"🎉 Success! {updated} staff entries update ho gayi hain.")
                                st.download_button(label="📥 Download Updated Attendance Excel", data=output, file_name=f"Updated_Attendance_Date_{date_val}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                            except Exception as e: st.error(f"Data save karte waqt error: {e}")